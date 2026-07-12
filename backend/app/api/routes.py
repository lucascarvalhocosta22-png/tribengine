import os
import io
import zipfile
from fastapi import APIRouter, UploadFile, File, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List
from datetime import datetime

from app.database import get_db
from app.models.nfe import NFe, ItemNFe
from app.models.simulacao import Simulacao, ResultadoSimulacao
from app.models.classificacao import HistoricoClassificacao
from app.schemas.nfe_schema import NFeResponse, ItemNFeResponse, SimulacaoRequest, SimulacaoResponse, AuditoriaResponse, ManualData
from app.services.xml_parser import parse_nfe_xml
from app.services.calculadora import calculadora
from app.services.auditoria import auditoria
from app.services.simples_nacional import is_simples_nacional, listar_fornecedores

router = APIRouter()


def _save_nfe_from_parsed(db: Session, dados: dict, xml_str: str, tipo_declarado: str = None) -> int:
    chave = dados.get("chave_acesso", "")
    existing = db.query(NFe).filter(NFe.chave_acesso == chave).first()

    remetente_cnpj = dados.get("remetente_cnpj", "")
    sn_status = "Nao"
    if tipo_declarado == "compra" and remetente_cnpj:
        try:
            sn_status = "Sim" if is_simples_nacional(remetente_cnpj) else "Nao"
        except Exception:
            sn_status = "Nao"

    if existing:
        nfe_id = existing.id
    else:
        nfe = NFe(
            chave_acesso=chave,
            tipo_declarado=tipo_declarado,
            simples_nacional=sn_status,
            numero=dados.get("numero", ""),
            serie=dados.get("serie", ""),
            data_emissao=datetime.fromisoformat(dados.get("data_emissao", "").replace("Z", "+00:00")) if dados.get("data_emissao") else None,
            data_recebimento=datetime.fromisoformat(dados.get("data_recebimento", "").replace("Z", "+00:00")) if dados.get("data_recebimento") else None,
            destinatario_nome=dados.get("destinatario_nome", ""),
            destinatario_cnpj=dados.get("destinatario_cnpj", ""),
            remetente_nome=dados.get("remetente_nome", ""),
            remetente_cnpj=dados.get("remetente_cnpj", ""),
            valor_total=dados.get("valor_total", 0),
            valor_base_calculo=dados.get("valor_base_calculo", 0),
            valor_icms=dados.get("valor_icms", 0),
            valor_frete=dados.get("valor_frete", 0),
            valor_seguro=dados.get("valor_seguro", 0),
            valor_despesas=dados.get("valor_despesas", 0),
            valor_desconto=dados.get("valor_desconto", 0),
            valor_total_tributos=dados.get("valor_total_tributos", 0),
            natureza_operacao=dados.get("natureza_operacao", ""),
            tipo_operacao=dados.get("tipo_operacao", ""),
            finalidade=dados.get("finalidade", ""),
            xml_original=xml_str,
        )
        db.add(nfe)
        db.flush()
        nfe_id = nfe.id

    if existing:
        db.query(ItemNFe).filter(ItemNFe.nfe_id == nfe_id).delete()
        if sn_status and existing.simples_nacional != sn_status:
            existing.simples_nacional = sn_status

    for item_data in dados.get("itens", []):
        item_data["simples_nacional"] = sn_status
        item_data["tipo_declarado"] = tipo_declarado
        calc = calculadora.calcular_item(item_data)
        auditoria_res = auditoria.auditar_item(item_data)
        item = ItemNFe(
            nfe_id=nfe_id,
            codigo_produto=item_data.get("codigo_produto", ""),
            descricao=item_data.get("descricao", ""),
            ncm=item_data.get("ncm", ""),
            cest=item_data.get("cest", ""),
            cfop=item_data.get("cfop", ""),
            uom=item_data.get("uom", "UN"),
            quantidade=item_data.get("quantidade", 1),
            valor_unitario=item_data.get("valor_unitario", 0),
            valor_total=item_data.get("valor_total", 0),
            gtin=item_data.get("gtin", ""),
            cclass_trib=item_data.get("cclass_trib", "01.99.99"),
            cst_ibms=item_data.get("cst_ibms", "00"),
            cst_cbs=item_data.get("cst_cbs", "00"),
            cst_is=item_data.get("cst_is", "00"),
            aliquota_ibms=calc["valor_ibms"] / calc["base_ibms"] * 100 if calc["base_ibms"] > 0 else item_data.get("aliquota_ibms", 0),
            aliquota_cbs=calc["valor_cbs"] / calc["base_cbs"] * 100 if calc["base_cbs"] > 0 else item_data.get("aliquota_cbs", 0),
            aliquota_is=calc["valor_is"] / calc["base_is"] * 100 if calc["base_is"] > 0 else item_data.get("aliquota_is", 0),
            base_ibms=calc["base_ibms"],
            base_cbs=calc["base_cbs"],
            base_is=calc["base_is"],
            valor_ibms=calc["valor_ibms"],
            valor_cbs=calc["valor_cbs"],
            valor_is=calc["valor_is"],
            incidencia_ibms="Sim" if calc["valor_ibms"] > 0 else "Nao",
            incidencia_cbs="Sim" if calc["valor_cbs"] > 0 else "Nao",
            incidencia_is="Sim" if calc["valor_is"] > 0 else "Nao",
            regime_especifico=item_data.get("regime_especifico", ""),
            regime_diferenciado=item_data.get("regime_diferenciado", ""),
            regime_favorecido=item_data.get("regime_favorecido", "Nao"),
            suspensao=item_data.get("suspensao", "Nao"),
            imunidade=item_data.get("imunidade", "Nao"),
            isencao=item_data.get("isencao", "Nao"),
            reducao_base=item_data.get("reducao_base", 0),
            reducao_aliquota=item_data.get("reducao_aliquota", 0),
            aliquota_zero=item_data.get("aliquota_zero", "Nao"),
            cashback=item_data.get("cashback", "Nao"),
            split_payment=item_data.get("split_payment", "Nao"),
            monofasico=item_data.get("monofasico", "Nao"),
            cesta_basica=item_data.get("cesta_basica", "Nao"),
            confianca_classificacao=item_data.get("confianca", 0),
            fundamentacao_legal=item_data.get("fundamentacao_legal", ""),
            historico_decisao=item_data.get("artigos_utilizados", []),
            precisa_revisao=item_data.get("precisa_revisao", "Sim"),
        )
        db.add(item)
        hist = HistoricoClassificacao(
            ncm=item_data.get("ncm", ""),
            cest=item_data.get("cest", ""),
            descricao=item_data.get("descricao", ""),
            cclass_trib=item_data.get("cclass_trib", ""),
            cst_ibms=item_data.get("cst_ibms", ""),
            cst_cbs=item_data.get("cst_cbs", ""),
            cst_is=item_data.get("cst_is", ""),
            confianca=item_data.get("confianca", 0),
            fundamentacao=item_data.get("fundamentacao_legal", ""),
            artigos_utilizados=item_data.get("artigos_utilizados", []),
            fonte="XML NF-e",
        )
        db.add(hist)
    return nfe_id


@router.post("/upload/xml", response_model=NFeResponse)
async def upload_xml(
    file: UploadFile = File(...),
    tipo: str = Query("auto", pattern="^(compra|venda|auto)$"),
    db: Session = Depends(get_db),
):
    try:
        content = await file.read()
        tipo_dec = tipo if tipo != "auto" else ""
        dados = parse_nfe_xml(content, tipo_dec)
        nfe_id = _save_nfe_from_parsed(db, dados, content.decode("utf-8", errors="replace"), tipo_dec or None)
        db.commit()
        nfe = db.query(NFe).filter(NFe.id == nfe_id).first()
        return _nfe_to_response(nfe)
    except Exception as e:
        db.rollback()
        raise HTTPException(500, detail=str(e)[:500])


@router.post("/upload/batch")
async def upload_batch(
    files: List[UploadFile] = File(...),
    tipo: str = Query("auto", pattern="^(compra|venda|auto)$"),
    db: Session = Depends(get_db),
):
    tipo_dec = tipo if tipo != "auto" else ""
    results = []
    for file in files:
        if not file.filename.endswith(".xml"):
            results.append({"arquivo": file.filename, "erro": "Nao e XML", "sucesso": False})
            continue
        try:
            content = await file.read()
            dados = parse_nfe_xml(content, tipo_dec)
            xml_str = content.decode("utf-8", errors="replace")
            nfe_id = _save_nfe_from_parsed(db, dados, xml_str, tipo_dec or None)
            results.append({
                "arquivo": file.filename, "chave": dados.get("chave_acesso", ""),
                "itens": len(dados.get("itens", [])), "nfe_id": nfe_id, "sucesso": True,
            })
        except Exception as e:
            results.append({"arquivo": file.filename, "erro": str(e)[:200], "sucesso": False})
    db.commit()
    s = sum(1 for r in results if r["sucesso"])
    e = sum(1 for r in results if not r["sucesso"])
    return {"mensagem": f"{s}/{len(results)} importados, {e} erro(s)", "resultados": results}

@router.post("/upload/zip")
async def upload_zip(
    file: UploadFile = File(...),
    tipo: str = Query("auto", pattern="^(compra|venda|auto)$"),
    db: Session = Depends(get_db),
):
    if not file.filename.endswith(".zip"):
        raise HTTPException(400, "Arquivo deve ser ZIP")

    content = await file.read()
    zip_path = os.path.join("uploads", file.filename)
    with open(zip_path, "wb") as f:
        f.write(content)

    results = []
    with zipfile.ZipFile(zip_path, "r") as z:
        for name in z.namelist():
            if name.endswith(".xml"):
                xml_bytes = z.read(name)
                try:
                    tipo_dec = tipo if tipo != "auto" else ""
                    dados = parse_nfe_xml(xml_bytes, tipo_dec)
                    nfe_id = _save_nfe_from_parsed(db, dados, xml_bytes.decode("utf-8", errors="replace"), tipo_dec or None)
                    results.append({
                        "arquivo": name,
                        "chave": dados.get("chave_acesso", ""),
                        "itens": len(dados.get("itens", [])),
                        "nfe_id": nfe_id,
                        "sucesso": True,
                    })
                except Exception as e:
                    results.append({
                        "arquivo": name,
                        "erro": str(e),
                        "sucesso": False,
                    })

    db.commit()
    return {"mensagem": f"Processados {len(results)} arquivos, salvos no banco", "resultados": results}


@router.get("/nfe")
async def listar_nfes(db: Session = Depends(get_db)):
    nfes = db.query(NFe).order_by(NFe.data_emissao.desc()).all()
    return [{
        "id": n.id, "chave_acesso": n.chave_acesso, "numero": n.numero,
        "serie": n.serie, "data_emissao": n.data_emissao.isoformat() if n.data_emissao else None,
        "destinatario_nome": n.destinatario_nome, "destinatario_cnpj": n.destinatario_cnpj,
        "remetente_nome": n.remetente_nome, "remetente_cnpj": n.remetente_cnpj,
        "valor_total": n.valor_total, "natureza_operacao": n.natureza_operacao,
        "tipo_declarado": n.tipo_declarado, "simples_nacional": n.simples_nacional,
    } for n in nfes]


@router.get("/nfe/sem-credito")
async def nfe_sem_credito(db: Session = Depends(get_db)):
    nfes = db.query(NFe).filter(
        NFe.simples_nacional == "Sim",
        NFe.tipo_declarado == "compra",
    ).all()
    return [{
        "id": n.id, "numero": n.numero, "serie": n.serie,
        "data_emissao": n.data_emissao.isoformat() if n.data_emissao else None,
        "remetente_nome": n.remetente_nome, "remetente_cnpj": n.remetente_cnpj,
        "valor_total": n.valor_total,
        "motivo": "Fornecedor optante pelo Simples Nacional — não gera crédito de IBS/CBS (LC 214/2025, Art. 28)",
    } for n in nfes]


@router.get("/nfe/{nfe_id}", response_model=NFeResponse)
async def obter_nfe(nfe_id: int, db: Session = Depends(get_db)):
    nfe = db.query(NFe).filter(NFe.id == nfe_id).first()
    if not nfe:
        raise HTTPException(404, "NF-e não encontrada")
    return _nfe_to_response(nfe)


@router.delete("/nfe")
async def deletar_todas_nfes(db: Session = Depends(get_db)):
    total_nfes = db.query(NFe).count()
    total_itens = db.query(ItemNFe).count()
    db.query(ItemNFe).delete()
    db.query(NFe).delete()
    db.commit()
    return {"mensagem": f"Deletadas {total_nfes} NF-e(s) e {total_itens} ite(ns)"}


@router.delete("/nfe/{nfe_id}")
async def deletar_nfe(nfe_id: int, db: Session = Depends(get_db)):
    nfe = db.query(NFe).filter(NFe.id == nfe_id).first()
    if not nfe:
        raise HTTPException(404, "NF-e não encontrada")
    db.delete(nfe)
    db.commit()
    return {"mensagem": "NF-e deletada com sucesso"}


@router.get("/nfe/{nfe_id}/auditoria", response_model=List[AuditoriaResponse])
async def auditar_nfe(nfe_id: int, db: Session = Depends(get_db)):
    itens = db.query(ItemNFe).filter(ItemNFe.nfe_id == nfe_id).all()
    resultados = []
    for item in itens:
        item_dict = {
            "ncm": item.ncm,
            "cest": item.cest,
            "cfop": item.cfop,
            "descricao": item.descricao,
            "cclass_trib": item.cclass_trib,
            "cst_ibms": item.cst_ibms,
            "cst_cbs": item.cst_cbs,
            "cst_is": item.cst_is,
            "aliquota_ibms": item.aliquota_ibms,
            "aliquota_cbs": item.aliquota_cbs,
            "aliquota_is": item.aliquota_is,
            "confianca": item.confianca_classificacao,
            "monofasico": item.monofasico,
            "aliquota_zero": item.aliquota_zero,
            "reducao_base": item.reducao_base,
            "imunidade": item.imunidade,
            "cashback": item.cashback,
        }
        audit = auditoria.auditar_item(item_dict)
        resultados.append(AuditoriaResponse(
            item_id=item.id,
            descricao=item.descricao,
            **audit,
        ))
    return resultados


@router.post("/simular", response_model=SimulacaoResponse)
async def simular(req: SimulacaoRequest, db: Session = Depends(get_db)):
    itens = []
    if req.manual:
        if req.manual.total_compras > 0:
            itens.append({
                "ncm": "00000000",
                "cfop": "1.001",
                "descricao": "Compras - Manual",
                "valor_total": req.manual.total_compras,
                "aliquota_ibms": 17.7,
                "aliquota_cbs": 8.8,
                "aliquota_is": 0,
                "reducao_base": 40 if req.manual.reducao_compras else 0,
                "aliquota_zero": "Nao",
                "imunidade": "Nao",
                "isencao": "Nao",
                "suspensao": "Nao",
                "monofasico": "Nao",
                "cesta_basica": "Nao",
                "cclass_trib": "01.99.99",
                "cest": "",
                "tipo_declarado": "compra",
            })
        if req.manual.total_vendas > 0:
            itens.append({
                "ncm": "00000000",
                "cfop": "5.001",
                "descricao": "Vendas - Manual",
                "valor_total": req.manual.total_vendas,
                "aliquota_ibms": 17.7,
                "aliquota_cbs": 8.8,
                "aliquota_is": 0,
                "reducao_base": 0,
                "aliquota_zero": "Nao",
                "imunidade": "Nao",
                "isencao": "Nao",
                "suspensao": "Nao",
                "monofasico": "Nao",
                "cesta_basica": "Nao",
                "cclass_trib": "01.02.01" if req.manual.reducao_vendas else "01.99.99",
                "cest": "",
                "tipo_declarado": "venda",
            })
    if req.nfe_ids:
        itens_query = db.query(ItemNFe, NFe.tipo_declarado, NFe.simples_nacional).join(NFe, ItemNFe.nfe_id == NFe.id).filter(ItemNFe.nfe_id.in_(req.nfe_ids)).all()
        for item, tipo_declarado, simples_nacional in itens_query:
            itens.append({
                "ncm": item.ncm, "cfop": item.cfop, "descricao": item.descricao,
                "valor_total": item.valor_total,
                "aliquota_ibms": item.aliquota_ibms, "aliquota_cbs": item.aliquota_cbs, "aliquota_is": item.aliquota_is,
                "reducao_base": item.reducao_base, "aliquota_zero": item.aliquota_zero,
                "imunidade": item.imunidade, "isencao": item.isencao, "suspensao": item.suspensao,
                "monofasico": item.monofasico, "cesta_basica": item.cesta_basica,
                "cclass_trib": item.cclass_trib, "cest": item.cest,
                "tipo_declarado": tipo_declarado, "simples_nacional": simples_nacional or "Nao",
            })

    simulacao = Simulacao(
        nome=req.nome,
        mes_inicio=req.mes_inicio,
        ano_inicio=req.ano_inicio,
        mes_fim=req.mes_fim,
        ano_fim=req.ano_fim,
        parametros={"nfe_ids": req.nfe_ids, "regime": req.regime},
    )
    db.add(simulacao)
    db.flush()

    resultados_list = []
    for mes in range(req.mes_inicio, req.mes_fim + 1):
        resultado = calculadora.simular_mensal(itens, mes, req.ano_inicio, regime=req.regime)
        res = ResultadoSimulacao(
            simulacao_id=simulacao.id,
            mes=mes,
            ano=req.ano_inicio,
            **{k: v for k, v in resultado.items() if k not in ["mes", "ano"]},
        )
        db.add(res)
        resultados_list.append(resultado)

    db.commit()
    return SimulacaoResponse(id=simulacao.id, nome=req.nome, resultados=resultados_list)


@router.get("/dashboard")
async def dashboard(db: Session = Depends(get_db)):
    from sqlalchemy import func, case, and_, or_

    total_nfes = db.query(NFe).count()
    total_itens = db.query(ItemNFe).count()

    compra = and_(NFe.tipo_declarado == "compra")
    venda = and_(NFe.tipo_declarado == "venda")
    sn = NFe.simples_nacional == "Sim"
    reducao = case((ItemNFe.cclass_trib.in_(["01.02.01", "200047"]), 0.6), else_=1.0)

    sums = db.query(
        func.sum(ItemNFe.valor_ibms * reducao).label("total_ibms"),
        func.sum(ItemNFe.valor_cbs * reducao).label("total_cbs"),
        func.sum(ItemNFe.valor_is).label("total_is"),
        func.sum(case((and_(compra, ~sn), ItemNFe.valor_ibms), else_=0)).label("ibs_credito"),
        func.sum(case((and_(compra, ~sn), ItemNFe.valor_cbs), else_=0)).label("cbs_credito"),
        func.sum(case((venda, ItemNFe.valor_ibms * reducao), else_=0)).label("ibs_debito"),
        func.sum(case((venda, ItemNFe.valor_cbs * reducao), else_=0)).label("cbs_debito"),
        func.sum(case((venda, ItemNFe.valor_is), else_=0)).label("is_debito"),
        func.sum(case((compra, ItemNFe.valor_total), else_=0)).label("total_compras"),
        func.sum(case((venda, ItemNFe.valor_total), else_=0)).label("total_vendas"),
        func.count(case((ItemNFe.reducao_base > 0, 1))).label("com_reducao"),
        func.count(case((or_(ItemNFe.imunidade == "Sim", ItemNFe.isencao == "Sim"), 1))).label("isentos"),
        func.count(case((ItemNFe.monofasico == "Sim", 1))).label("monofasicos"),
        func.count(case((ItemNFe.valor_is > 0, 1))).label("seletivos"),
        func.count(case((ItemNFe.precisa_revisao == "Sim", 1))).label("pendentes"),
    ).select_from(ItemNFe).join(NFe).first()

    total_ibms = sums.total_ibms or 0
    total_cbs = sums.total_cbs or 0
    total_is = sums.total_is or 0
    ibs_credito = sums.ibs_credito or 0
    cbs_credito = sums.cbs_credito or 0
    ibs_debito = sums.ibs_debito or 0
    cbs_debito = sums.cbs_debito or 0
    is_debito = sums.is_debito or 0
    total_compras = sums.total_compras or 0
    total_vendas = sums.total_vendas or 0
    com_reducao = sums.com_reducao or 0
    isentos = sums.isentos or 0
    monofasicos = sums.monofasicos or 0
    seletivos = sums.seletivos or 0
    pendentes = sums.pendentes or 0

    fornecedores_sn = db.query(NFe.remetente_cnpj).filter(sn, compra).distinct().count()
    compras_sem_credito = db.query(ItemNFe).join(NFe).filter(sn, compra).count()

    motivos_data = db.query(
        case(
            (sn, "Simples Nacional"),
            (ItemNFe.cesta_basica == "Sim", "Cesta Básica"),
            (ItemNFe.aliquota_zero == "Sim", "Alíquota Zero"),
            (ItemNFe.imunidade == "Sim", "Imunidade"),
            (ItemNFe.isencao == "Sim", "Isenção"),
            (ItemNFe.monofasico == "Sim", "Monofásico"),
            (ItemNFe.suspensao == "Sim", "Suspensão"),
            else_="Alíquota Efetiva Zero",
        ).label("motivo"),
        func.count(ItemNFe.id).label("itens"),
        func.sum(ItemNFe.valor_total).label("valor"),
    ).join(NFe).filter(compra, ItemNFe.valor_ibms == 0, ItemNFe.valor_cbs == 0).group_by("motivo").all()

    sem_credito_motivos = [{"motivo": m.motivo, "itens": m.itens, "valor": round(m.valor or 0, 2)} for m in motivos_data]
    compras_sem_credito_total = sum(m["itens"] for m in sem_credito_motivos)
    compras_sem_credito_valor = round(sum(m["valor"] for m in sem_credito_motivos), 2)

    total_creditos = ibs_credito + cbs_credito
    total_debitos = ibs_debito + cbs_debito + is_debito

    carga_tributaria = total_ibms + total_cbs + total_is
    receita_total = total_vendas if total_vendas > 0 else 1
    carga_percentual = round((carga_tributaria / receita_total * 100), 2)
    saldo_liquido = total_creditos - total_debitos
    saldo_credor = max(0, saldo_liquido)
    saldo_devedor = max(0, -saldo_liquido)

    return {
        "total_nfes": total_nfes,
        "total_itens": total_itens,
        "total_compras": round(total_compras, 2),
        "total_vendas": round(total_vendas, 2),
        "total_ibms": round(total_ibms, 2),
        "total_cbs": round(total_cbs, 2),
        "total_is": round(total_is, 2),
        "ibs_credito": round(ibs_credito, 2),
        "cbs_credito": round(cbs_credito, 2),
        "ibs_debito": round(ibs_debito, 2),
        "cbs_debito": round(cbs_debito, 2),
        "is_debito": round(is_debito, 2),
        "total_creditos": round(total_creditos, 2),
        "total_debitos": round(total_debitos, 2),
        "saldo_liquido": round(saldo_liquido, 2),
        "saldo_credor": round(saldo_credor, 2),
        "saldo_devedor": round(saldo_devedor, 2),
        "carga_tributaria": round(carga_tributaria, 2),
        "carga_percentual": carga_percentual,
        "com_reducao": com_reducao,
        "isentos": isentos,
        "monofasicos": monofasicos,
        "seletivos": seletivos,
        "pendentes": pendentes,
        "fornecedores_simples": fornecedores_sn,
        "compras_sem_credito_sn": compras_sem_credito,
        "compras_sem_credito_total": compras_sem_credito_total,
        "compras_sem_credito_valor": compras_sem_credito_valor,
        "compras_sem_credito_motivos": sorted(sem_credito_motivos, key=lambda x: -x["itens"]),
    }


@router.get("/classificar")
async def classificar_manual(
    ncm: str = Query(""),
    descricao: str = Query(""),
    cfop: str = Query(""),
    db: Session = Depends(get_db),
):
    from app.engine.classificador import engine as trib_engine
    resultado = trib_engine.classificar(ncm=ncm, cest="", cfop=cfop, descricao=descricao)
    calc = calculadora.calcular_item(resultado)
    audit = auditoria.auditar_item(resultado)
    return {
        "classificacao": resultado,
        "calculo": calc,
        "auditoria": audit,
    }


@router.post("/dashboard/simular", response_model=SimulacaoResponse)
async def simular_tudo(db: Session = Depends(get_db), regime: str = "geral"):
    todos_ids = [n.id for n in db.query(NFe.id).all()]
    if not todos_ids:
        return SimulacaoResponse(id=0, nome="Simulação Automática", resultados=[])
    req = SimulacaoRequest(nome="Simulação Automática", nfe_ids=todos_ids, regime=regime)
    return await simular(req, db)


def _nfe_to_response(nfe: NFe) -> dict:
    itens = []
    for item in nfe.itens:
        calc = calculadora.calcular_item({
            "valor_total": item.valor_total,
            "aliquota_ibms": item.aliquota_ibms,
            "aliquota_cbs": item.aliquota_cbs,
            "aliquota_is": item.aliquota_is,
            "reducao_base": item.reducao_base,
            "aliquota_zero": item.aliquota_zero,
            "imunidade": item.imunidade,
            "isencao": item.isencao,
            "suspensao": item.suspensao,
            "monofasico": item.monofasico,
            "cesta_basica": item.cesta_basica,
        })
        itens.append({
            "id": item.id,
            "codigo_produto": item.codigo_produto,
            "descricao": item.descricao,
            "ncm": item.ncm,
            "cest": item.cest,
            "cfop": item.cfop,
            "uom": item.uom,
            "quantidade": item.quantidade,
            "valor_unitario": item.valor_unitario,
            "valor_total": item.valor_total,
            "gtin": item.gtin,
            "cclass_trib": item.cclass_trib,
            "cclasstrib_descricao": "",
            "cst_ibms": item.cst_ibms,
            "cst_cbs": item.cst_cbs,
            "cst_is": item.cst_is,
            "aliquota_ibms": item.aliquota_ibms,
            "aliquota_cbs": item.aliquota_cbs,
            "aliquota_is": item.aliquota_is,
            "base_ibms": calc["base_ibms"],
            "base_cbs": calc["base_cbs"],
            "base_is": calc["base_is"],
            "valor_ibms": calc["valor_ibms"],
            "valor_cbs": calc["valor_cbs"],
            "valor_is": calc["valor_is"],
            "carga_total": calc["carga_total"],
            "regime_especifico": item.regime_especifico,
            "regime_diferenciado": item.regime_diferenciado,
            "regime_favorecido": item.regime_favorecido,
            "suspensao": item.suspensao,
            "imunidade": item.imunidade,
            "isencao": item.isencao,
            "reducao_base": item.reducao_base,
            "reducao_aliquota": item.reducao_aliquota,
            "aliquota_zero": item.aliquota_zero,
            "cashback": item.cashback,
            "split_payment": item.split_payment,
            "monofasico": item.monofasico,
            "cesta_basica": item.cesta_basica,
            "confianca": item.confianca_classificacao,
            "fundamentacao_legal": item.fundamentacao_legal,
            "artigos_utilizados": item.historico_decisao,
            "precisa_revisao": item.precisa_revisao,
        })
    return {
        "id": nfe.id,
        "chave_acesso": nfe.chave_acesso,
        "numero": nfe.numero,
        "serie": nfe.serie,
        "data_emissao": nfe.data_emissao.isoformat() if nfe.data_emissao else None,
        "destinatario_nome": nfe.destinatario_nome,
        "destinatario_cnpj": nfe.destinatario_cnpj,
        "remetente_nome": nfe.remetente_nome,
        "remetente_cnpj": nfe.remetente_cnpj,
        "valor_total": nfe.valor_total,
        "natureza_operacao": nfe.natureza_operacao,
        "tipo_declarado": nfe.tipo_declarado,
        "simples_nacional": nfe.simples_nacional,
        "itens": itens,
    }


@router.get("/credito/analise")
async def credito_analise(db: Session = Depends(get_db)):
    itens = db.query(ItemNFe, NFe).join(NFe, ItemNFe.nfe_id == NFe.id).filter(NFe.tipo_declarado == "compra").all()
    fornecedores = {}
    for item, nfe in itens:
        cnpj = nfe.remetente_cnpj or "0000"
        if cnpj not in fornecedores:
            fornecedores[cnpj] = {
                "cnpj": cnpj, "fornecedor": nfe.remetente_nome or "",
                "total_compras": 0, "total_itens": 0,
                "credito_ibms": 0, "credito_cbs": 0,
                "itens": [],
            }
        f = fornecedores[cnpj]
        f["total_compras"] += item.valor_total or 0
        f["total_itens"] += 1

        motivos = []
        if nfe.simples_nacional == "Sim":
            motivos.append({"motivo": "Simples Nacional", "detalhe": "Fornecedor SN não gera crédito (LC 214/2025, Art. 28)"})
        if item.aliquota_zero == "Sim":
            motivos.append({"motivo": "Alíquota Zero", "detalhe": "Produto com alíquota zero de IBS/CBS"})
        if item.imunidade == "Sim":
            motivos.append({"motivo": "Imunidade", "detalhe": "Produto imune a IBS/CBS"})
        if item.isencao == "Sim":
            motivos.append({"motivo": "Isenção", "detalhe": "Produto isento de IBS/CBS"})
        if item.monofasico == "Sim":
            motivos.append({"motivo": "Monofásico", "detalhe": "Produto monofásico — sem crédito"})
        if item.suspensao == "Sim":
            motivos.append({"motivo": "Suspensão", "detalhe": "Suspensão de IBS/CBS — sem crédito"})
        if item.cesta_basica == "Sim":
            motivos.append({"motivo": "Cesta Básica", "detalhe": "Cesta básica — alíquota zero"})

        gerou_credito = (item.valor_ibms or 0) > 0 or (item.valor_cbs or 0) > 0
        if not motivos and not gerou_credito:
            motivos.append({"motivo": "Alíquota Efetiva Zero", "detalhe": "Alíquota de IBS/CBS resulta em R$ 0,00"})

        f["credito_ibms"] += item.valor_ibms or 0
        f["credito_cbs"] += item.valor_cbs or 0
        f["itens"].append({
            "id": item.id, "descricao": item.descricao, "ncm": item.ncm,
            "cclass_trib": item.cclass_trib, "cfop": item.cfop,
            "valor_total": item.valor_total,
            "valor_ibms": item.valor_ibms, "valor_cbs": item.valor_cbs,
            "gerou_credito": gerou_credito,
            "motivos_sem_credito": motivos,
            "aliquota_ibms": item.aliquota_ibms, "aliquota_cbs": item.aliquota_cbs,
            "aliquota_zero": item.aliquota_zero, "imunidade": item.imunidade,
            "isencao": item.isencao, "monofasico": item.monofasico,
            "suspensao": item.suspensao, "cesta_basica": item.cesta_basica,
        })

    for f in fornecedores.values():
        f["total_credito"] = f["credito_ibms"] + f["credito_cbs"]
        f["itens_sem_credito"] = sum(1 for i in f["itens"] if not i["gerou_credito"])
        f["itens_com_credito"] = sum(1 for i in f["itens"] if i["gerou_credito"])

    return {
        "total_fornecedores": len(fornecedores),
        "total_itens": len(itens),
        "total_credito_ibms": round(sum(f["credito_ibms"] for f in fornecedores.values()), 2),
        "total_credito_cbs": round(sum(f["credito_cbs"] for f in fornecedores.values()), 2),
        "fornecedores": sorted(fornecedores.values(), key=lambda x: -x["total_compras"]),
    }


@router.get("/credito/export/excel")
async def credito_export_excel(db: Session = Depends(get_db)):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side
    except ImportError:
        from fastapi.responses import JSONResponse
        return JSONResponse({"detail": "openpyxl não instalado"}, status_code=500)

    import io
    from fastapi.responses import StreamingResponse

    itens = db.query(ItemNFe, NFe).join(NFe, ItemNFe.nfe_id == NFe.id).filter(NFe.tipo_declarado == "compra").all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Análise de Créditos"

    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    tb = Border(left=Side(style='thin', color='475569'), right=Side(style='thin', color='475569'),
                top=Side(style='thin', color='475569'), bottom=Side(style='thin', color='475569'))

    headers = ["Fornecedor", "CNPJ", "NCM", "Descrição", "cClassTrib", "CFOP",
               "Valor Total", "IBS", "CBS", "Gerou Crédito?", "Motivo"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hfill; c.border = tb

    row = 2
    for item, nfe in itens:
        gerou = (item.valor_ibms or 0) > 0 or (item.valor_cbs or 0) > 0
        motivos = []
        if nfe.simples_nacional == "Sim": motivos.append("SN")
        if item.aliquota_zero == "Sim": motivos.append("Aliq Zero")
        if item.imunidade == "Sim": motivos.append("Imunidade")
        if item.isencao == "Sim": motivos.append("Isenção")
        if item.monofasico == "Sim": motivos.append("Monofásico")
        if item.suspensao == "Sim": motivos.append("Suspensão")
        if item.cesta_basica == "Sim": motivos.append("Cesta Básica")
        if not gerou and not motivos: motivos.append("Alíquota Zero Efetiva")

        vals = [nfe.remetente_nome, nfe.remetente_cnpj, item.ncm, item.descricao,
                item.cclass_trib, item.cfop, item.valor_total, item.valor_ibms,
                item.valor_cbs, "Sim" if gerou else "Não", "; ".join(motivos)]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = tb
        row += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 14
    ws.column_dimensions['K'].width = 30

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=analise_credito.xlsx"})


@router.get("/comparativo/old-new")
async def comparativo_old_new(db: Session = Depends(get_db)):
    itens = db.query(ItemNFe, NFe).join(NFe, ItemNFe.nfe_id == NFe.id).filter(NFe.tipo_declarado == "venda").all()
    itens_compra = db.query(ItemNFe, NFe).join(NFe, ItemNFe.nfe_id == NFe.id).filter(NFe.tipo_declarado == "compra").all()
    nfe_sn = {n.id: n.simples_nacional or "Nao" for n in db.query(NFe).all()}

    total_5102 = sum(i.valor_total for i, _ in itens if i.cfop == "5102")
    total_5405 = sum(i.valor_total for i, _ in itens if i.cfop == "5405")
    outros = sum(i.valor_total for i, _ in itens if i.cfop not in ("5102", "5405"))

    def _fr(cclass): return 0.6 if cclass in ("01.02.01", "200047") else 1.0

    ibs_5102 = sum((i.valor_ibms or 0) * _fr(i.cclass_trib) for i, _ in itens if i.cfop == "5102")
    cbs_5102 = sum((i.valor_cbs or 0) * _fr(i.cclass_trib) for i, _ in itens if i.cfop == "5102")
    ibs_5405 = sum((i.valor_ibms or 0) * _fr(i.cclass_trib) for i, _ in itens if i.cfop == "5405")
    cbs_5405 = sum((i.valor_cbs or 0) * _fr(i.cclass_trib) for i, _ in itens if i.cfop == "5405")
    ibs_outros = sum((i.valor_ibms or 0) * _fr(i.cclass_trib) for i, _ in itens if i.cfop not in ("5102", "5405"))
    cbs_outros = sum((i.valor_cbs or 0) * _fr(i.cclass_trib) for i, _ in itens if i.cfop not in ("5102", "5405"))

    # Old system: ICMS 4.15% on 5102, 0% on 5405; PIS/COFINS 4.65% on 5102, 0% on 5405
    icms_5102 = round(total_5102 * 0.0415, 2)
    icms_5405 = 0
    icms_outros = round(outros * 0.0415, 2)
    pis_cofins_5102 = round(total_5102 * 0.0465, 2)
    pis_cofins_5405 = 0
    pis_cofins_outros = round(outros * 0.0465, 2)

    novo_5102 = round(ibs_5102 + cbs_5102, 2)
    novo_5405 = round(ibs_5405 + cbs_5405, 2)
    novo_outros = round(ibs_outros + cbs_outros, 2)

    def _antigo_total(icms, pc): return round(icms + pc, 2)
    antigo_5102 = _antigo_total(icms_5102, pis_cofins_5102)
    antigo_5405 = _antigo_total(icms_5405, pis_cofins_5405)
    antigo_outros = _antigo_total(icms_outros, pis_cofins_outros)

    return {
        "cfop_5102": {
            "total": round(total_5102, 2),
            "itens": sum(1 for i, _ in itens if i.cfop == "5102"),
            "antigo_icms": icms_5102,
            "antigo_pis_cofins": pis_cofins_5102,
            "antigo_total": antigo_5102,
            "novo_ibs": round(ibs_5102, 2),
            "novo_cbs": round(cbs_5102, 2),
            "novo_total": novo_5102,
            "diferenca": round(novo_5102 - antigo_5102, 2),
            "variacao_percentual": round(((novo_5102 - antigo_5102) / antigo_5102 * 100) if antigo_5102 else 0, 2),
        },
        "cfop_5405": {
            "total": round(total_5405, 2),
            "itens": sum(1 for i, _ in itens if i.cfop == "5405"),
            "antigo_icms": icms_5405,
            "antigo_pis_cofins": pis_cofins_5405,
            "antigo_total": antigo_5405,
            "novo_ibs": round(ibs_5405, 2),
            "novo_cbs": round(cbs_5405, 2),
            "novo_total": novo_5405,
            "diferenca": round(novo_5405 - antigo_5405, 2),
            "variacao_percentual": round(((novo_5405 - antigo_5405) / max(1, antigo_5405)) * 100, 2),
        },
        "outros_cfop": {
            "total": round(outros, 2),
            "itens": sum(1 for i, _ in itens if i.cfop not in ("5102", "5405")),
            "antigo_icms": icms_outros,
            "antigo_pis_cofins": pis_cofins_outros,
            "antigo_total": antigo_outros,
            "novo_ibs": round(ibs_outros, 2),
            "novo_cbs": round(cbs_outros, 2),
            "novo_total": novo_outros,
            "diferenca": round(novo_outros - antigo_outros, 2),
            "variacao_percentual": round(((novo_outros - antigo_outros) / max(1, antigo_outros)) * 100, 2),
        },
        "total_geral": {
            "total_vendas": round(total_5102 + total_5405 + outros, 2),
            "total_itens": len(itens),
            "antigo_total": round(antigo_5102 + antigo_5405 + antigo_outros, 2),
            "novo_total": round(novo_5102 + novo_5405 + novo_outros, 2),
            "diferenca_total": round(novo_5102 + novo_5405 + novo_outros - antigo_5102 - antigo_5405 - antigo_outros, 2),
        },
        "creditos_compra": {
            "ibs_credito": round(sum(i.valor_ibms or 0 for i, n in itens_compra if nfe_sn.get(n.id, "Nao") != "Sim"), 2),
            "cbs_credito": round(sum(i.valor_cbs or 0 for i, n in itens_compra if nfe_sn.get(n.id, "Nao") != "Sim"), 2),
            "sem_credito_sn": round(sum(i.valor_ibms or 0 for i, n in itens_compra if nfe_sn.get(n.id, "Nao") == "Sim"), 2),
        },
    }


@router.get("/export/excel")
async def export_excel(db: Session = Depends(get_db)):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        from fastapi.responses import JSONResponse
        return JSONResponse({"detail": "openpyxl não instalado. Execute: pip install openpyxl"}, status_code=500)

    from app.engine.classificador import engine as trib_engine
    from app.services.calculadora import calculadora

    import io
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()

    # --- Sheet 1: Resumo ---
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    dash_data = {}
    itens = db.query(ItemNFe, NFe.tipo_declarado, NFe.simples_nacional).join(NFe, ItemNFe.nfe_id == NFe.id).all()
    def _ic(i, td): return (td == "compra") if td else (i.cfop[:1] in ["1", "2", "3"] if i.cfop else False)
    def _iv(i, td): return (td == "venda") if td else (i.cfop[:1] in ["5", "6", "7"] if i.cfop else False)
    dash_data["ibs_credito"] = sum(i.valor_ibms for i, td, sn in itens if _ic(i, td) and sn != "Sim")
    dash_data["cbs_credito"] = sum(i.valor_cbs for i, td, sn in itens if _ic(i, td) and sn != "Sim")
    dash_data["ibs_debito"] = sum(i.valor_ibms for i, td, _ in itens if _iv(i, td))
    dash_data["cbs_debito"] = sum(i.valor_cbs for i, td, _ in itens if _iv(i, td))
    dash_data["is_debito"] = sum(i.valor_is for i, td, _ in itens if _iv(i, td))
    dash_data["total_compras"] = sum(i.valor_total for i, td, _ in itens if _ic(i, td))
    dash_data["total_vendas"] = sum(i.valor_total for i, td, _ in itens if _iv(i, td))

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='475569'),
        right=Side(style='thin', color='475569'),
        top=Side(style='thin', color='475569'),
        bottom=Side(style='thin', color='475569'),
    )

    resumo_data = [
        ("Indicador", "Valor"),
        ("Total Compras", dash_data["total_compras"]),
        ("Total Vendas", dash_data["total_vendas"]),
        ("Crédito IBS", dash_data["ibs_credito"]),
        ("Crédito CBS", dash_data["cbs_credito"]),
        ("Débito IBS", dash_data["ibs_debito"]),
        ("Débito CBS", dash_data["cbs_debito"]),
        ("Débito IS", dash_data["is_debito"]),
        ("Saldo IBS", dash_data["ibs_credito"] - dash_data["ibs_debito"]),
        ("Saldo CBS", dash_data["cbs_credito"] - dash_data["cbs_debito"]),
        ("Saldo Líquido", dash_data["ibs_credito"] + dash_data["cbs_credito"] - dash_data["ibs_debito"] - dash_data["cbs_debito"] - dash_data["is_debito"]),
    ]
    for row_idx, (label, val) in enumerate(resumo_data, 1):
        cell_label = ws_resumo.cell(row=row_idx, column=1, value=label)
        cell_val = ws_resumo.cell(row=row_idx, column=2, value=round(val, 2) if isinstance(val, (int, float)) else val)
        if row_idx == 1:
            cell_label.font = header_font; cell_label.fill = header_fill
            cell_val.font = header_font; cell_val.fill = header_fill
        cell_label.border = thin_border; cell_val.border = thin_border
    ws_resumo.column_dimensions['A'].width = 20
    ws_resumo.column_dimensions['B'].width = 18

    # --- Sheet 2: Itens ---
    ws_itens = wb.create_sheet("Itens")
    headers = ["NF-e ID", "Tipo", "Remetente", "CNPJ Remetente", "Fornecedor SN",
               "Código", "Descrição", "NCM", "cClassTrib", "CFOP",
               "Valor Total", "Aliq IBS", "Aliq CBS", "Aliq IS",
               "Redução Base", "Base IBS", "Base CBS", "Base IS",
               "Valor IBS", "Valor CBS", "Valor IS", "Carga Total",
               "Cesta Básica", "Alíquota Zero", "Imunidade", "Isenção", "Suspensão", "Monofásico"]
    for col, h in enumerate(headers, 1):
        cell = ws_itens.cell(row=1, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border

    for row_idx, (item, td, sn) in enumerate(itens, 2):
        nfe = db.query(NFe).filter(NFe.id == item.nfe_id).first()
        vals = [
            item.nfe_id, td or "", nfe.remetente_nome if nfe else "", nfe.remetente_cnpj if nfe else "", sn or "Nao",
            item.codigo_produto, item.descricao, item.ncm, item.cclass_trib, item.cfop,
            item.valor_total, item.aliquota_ibms, item.aliquota_cbs, item.aliquota_is,
            item.reducao_base, item.base_ibms, item.base_cbs, item.base_is,
            item.valor_ibms, item.valor_cbs, item.valor_is,
            (item.valor_ibms or 0) + (item.valor_cbs or 0) + (item.valor_is or 0),
            item.cesta_basica, item.aliquota_zero, item.imunidade, item.isencao, item.suspensao, item.monofasico,
        ]
        for col, v in enumerate(vals, 1):
            cell = ws_itens.cell(row=row_idx, column=col, value=v)
            cell.border = thin_border

    # --- Sheet 3: NF-es sem crédito (SN) ---
    ws_sn = wb.create_sheet("Compras SN (sem crédito)")
    sn_headers = ["NF-e ID", "Número", "Série", "Data", "Fornecedor", "CNPJ", "Valor Total", "Motivo"]
    for col, h in enumerate(sn_headers, 1):
        cell = ws_sn.cell(row=1, column=col, value=h)
        cell.font = header_font; cell.fill = header_fill; cell.border = thin_border
    sn_nfes = db.query(NFe).filter(NFe.simples_nacional == "Sim", NFe.tipo_declarado == "compra").all()
    for row_idx, n in enumerate(sn_nfes, 2):
        vals = [
            n.id, n.numero, n.serie,
            n.data_emissao.isoformat() if n.data_emissao else "",
            n.remetente_nome, n.remetente_cnpj, n.valor_total,
            "Fornecedor Simples Nacional — sem crédito IBS/CBS (LC 214/2025)",
        ]
        for col, v in enumerate(vals, 1):
            cell = ws_sn.cell(row=row_idx, column=col, value=v)
            cell.border = thin_border

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=tributos.xlsx"})
